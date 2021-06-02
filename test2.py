import math
import numpy as nm
from scipy import integrate
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
wb.create_sheet(index = 1,title ='sheet2')

for j in range(1,11):
    for i in range(1,11) :
        y_1 =lambda x: 1/(math.sqrt(2*math.pi))*math.exp(-x**2/2)
        iy = integrate.quad(y_1,-nm.infty,1+(j/10))
        y_2 = 1/(1+math.exp(-1*(1+ (i/10))*(1+(j/1000))))
        ws.cell(row = i,column = j).value = y_2
        ws.cell(row = i,column = j).number_format = '0.0000000'
        ws.cell(row = i + 11,column = j).value,ws.cell(row = i + 12,column = j).value = iy
        ws.cell(row = i + 11,column = j).number_format = '0.0000000'
wb.save('test.xlsx')
